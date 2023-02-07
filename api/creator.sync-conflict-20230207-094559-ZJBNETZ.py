import openpyxl
from openpyxl.styles import PatternFill

RED="FF0000"

def otoczka_generator(data,sheet):
    for cell in sheet['A4':'BK4']:
        cell.fill=PatternFill(start_color=RED,end_color=RED,fill_type="solid")



def create_exel_file(NAME:str,len_week:int,PATH:str):
    wb=openpyxl.Workbook()
    zal=wb.active
    zal.title="Założenia"
    sheets=[]
    for i in range(len_week):
        sheets.append(wb.create_sheet(f"{i+1} Tydzien"))
    wb.save(f"{PATH}/{NAME}.xlsx")


create_exel_file("test","../output")
