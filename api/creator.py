import openpyxl
from openpyxl.styles import PatternFill

RED="FF0000"

def otoczka_generator(data,sheet):
    for row in sheet['A4':'BK4']:
        for cell in row:
            cell.fill=PatternFill(start_color=RED,end_color=RED,fill_type="solid")



def create_exel_file(NAME:str,len_week:int,PATH:str):
    wb=openpyxl.Workbook()
    zal=wb.active
    zal.title="Założenia"
    sheets=[]
    for i in range(len_week):
        sheets.append(wb.create_sheet(f"{i+1} Tydzien"))
    
    for i in sheets:
        otoczka_generator('',i)
    
    
    wb.save(f"{NAME}.xlsx")


create_exel_file("test2",5,"")
