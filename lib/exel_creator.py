import openpyxl

data={
  "name": "priorytet barki",
  "trener_id": 1,
  "user_id": 1,
  "ends": "",
  "trenings":
  [
      [        
              [1, "8-10", [[3, 2, 1, 1], [3, 3, 3, 3], [2, 2, 2, 2], [3, 3, 3, 3], [2, 2, 2, 2]]],
              [2, "6-8", [[3, 2, 1, 1], [3, 3, 3, 3], [2, 2, 2, 2], [3, 3, 3, 3], [2, 2, 2, 2]]],
              [4, "4-6", [[3, 2, 1, 1], [3, 3, 3, 3], [2, 2, 2, 2], [3, 3, 3, 3], [2, 2, 2, 2]]]
      ],
      [       [1, "8-10", [[3, 2, 1, 1], [3, 3, 3, 3], [2, 2, 2, 2], [3, 3, 3, 3], [2, 2, 2, 2]]],
              [2, "6-8", [[3, 2, 1, 1], [3, 3, 3, 3], [2, 2, 2, 2], [3, 3, 3, 3], [2, 2, 2, 2]]],
              [4, "4-6", [[3, 2, 1, 1], [3, 3, 3, 3], [2, 2, 2, 2], [3, 3, 3, 3], [2, 2, 2, 2]]]
      ]
  ]
}

trenings=data["trenings"]
print(trenings)

wb=openpyxl.Workbook()

zal=wb.active
zal.title="Założenia"
row=7

for trening in trenings:
    for i in trening:
        zal.cell(row=row,column=2,value=i[0])
        zal.cell(row=row,column=3,value=len(i[2]))
        zal.cell(row=row,column=4,value=i[1])
        cl=5
        for x in i[2]:
            pws=0
            for y in x:
                zal.cell(row=row,column=cl,value=y)
                cl+=1
            zal.cell(row=row,column=cl,value=pws)
            cl+=1